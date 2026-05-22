"""``/pages`` namespace commands for the AMX interactive CLI.

Mirrors the ``/lineage`` shape: a single Click group ``pages`` with
subcommands (``new``, ``list``, ``show``, ``edit``, ``export``,
``delete``). Inside the ``/pages`` REPL tab the user types ``/new``,
``/list``, etc.; from any other tab the same commands stay reachable
as ``/pages new`` / ``/pages list`` thanks to the cross-namespace
dispatch in :mod:`amx.cli_support.session`.

A bare subcommand call runs the wizard; flags are optional power-user
shortcuts.
"""

from __future__ import annotations

import sys
from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import click
from rich.console import Console
from rich.table import Table

from amx.config import AMXConfig
from amx.pages.factory import build_pages_service
from amx.pages.intent_templates import (
    INTENT_TEMPLATES,
    IntentTemplate,
    template_by_slug,
)
from amx.pages.intent_templates import (
    render as render_intent,
)
from amx.pages.service import PagesService
from amx.pages.types import AssetRef, SourceRef
from amx.utils.console import error, info, success, warn

FinalizeScope = Callable[..., Any]

_ASSET_KINDS: frozenset[str] = frozenset(
    {
        "db_profile",
        "db_database",
        "db_schema",
        "db_table",
        "db_column",
        "doc_profile",
        "lineage_artifact",
        "asset_notebook",
        "asset_job",
        "asset_pipeline",
        "asset_query",
        "asset_stream",
        "asset_streamlit",
    }
)

_REMOTE_ASSET_TABLES: dict[str, tuple[str, str]] = {
    "asset_notebook": (
        "remote_notebooks",
        "SELECT id, name, workspace_path, qualified_name "
        "FROM remote_notebooks WHERE profile_name = ? "
        "ORDER BY name LIMIT 200",
    ),
    "asset_job": (
        "remote_jobs",
        "SELECT id, name, '' AS workspace_path, '' AS qualified_name "
        "FROM remote_jobs WHERE profile_name = ? "
        "ORDER BY name LIMIT 200",
    ),
    "asset_pipeline": (
        "remote_pipelines",
        "SELECT id, name, '' AS workspace_path, target_schema AS qualified_name "
        "FROM remote_pipelines WHERE profile_name = ? "
        "ORDER BY name LIMIT 200",
    ),
    "asset_query": (
        "remote_queries",
        "SELECT id, name, kind AS workspace_path, warehouse AS qualified_name "
        "FROM remote_queries WHERE profile_name = ? "
        "ORDER BY executed_at DESC LIMIT 200",
    ),
    "asset_stream": (
        "remote_streams",
        "SELECT id, qualified_name AS name, source_table_fqn AS workspace_path, "
        "mode AS qualified_name FROM remote_streams WHERE profile_name = ? "
        "ORDER BY qualified_name LIMIT 200",
    ),
    "asset_streamlit": (
        "remote_streamlit_apps",
        "SELECT id, qualified_name AS name, main_file AS workspace_path, "
        "query_warehouse AS qualified_name FROM remote_streamlit_apps "
        "WHERE profile_name = ? ORDER BY qualified_name LIMIT 200",
    ),
}


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _svc(cfg: AMXConfig) -> PagesService:
    return build_pages_service(cfg)


def _parse_asset_flag(raw: str) -> AssetRef:
    if ":" not in raw:
        raise click.BadParameter(
            f"asset must be 'KIND:REF' (got {raw!r}). "
            f"Valid kinds: {', '.join(sorted(_ASSET_KINDS))}."
        )
    kind, ref = raw.split(":", 1)
    kind = kind.strip()
    ref = ref.strip()
    if kind not in _ASSET_KINDS:
        raise click.BadParameter(
            f"unknown asset kind {kind!r}. Valid: {', '.join(sorted(_ASSET_KINDS))}."
        )
    if not ref:
        raise click.BadParameter(f"asset {raw!r} has an empty ref")
    return AssetRef(kind=kind, ref=ref)  # type: ignore[arg-type]


def _pick_intent_template() -> IntentTemplate | None:
    """Render the picker and return the chosen template (None = custom)."""
    info("Pick an intent template:")
    for idx, tpl in enumerate(INTENT_TEMPLATES, start=1):
        click.echo(f"  {idx}. {tpl.label}")
    click.echo(f"  {len(INTENT_TEMPLATES) + 1}. Custom (free text)")
    raw = click.prompt(
        "Selection",
        default=str(len(INTENT_TEMPLATES) + 1),
        show_default=True,
    ).strip()
    try:
        choice = int(raw)
    except ValueError:
        return None
    if 1 <= choice <= len(INTENT_TEMPLATES):
        return INTENT_TEMPLATES[choice - 1]
    return None


def _collect_template_params(
    cfg: AMXConfig,
    template: IntentTemplate,
) -> tuple[list[AssetRef], dict[str, str]]:
    """Prompt for the assets the chosen template requires.

    Returns the asset list to attach and the placeholder map used to
    render the prompt skeleton into a concrete intent string.
    """
    assets: list[AssetRef] = []
    params: dict[str, str] = {}
    req = template.required_assets

    db_names = sorted((cfg.db_profiles or {}).keys())

    if req in {"one_db_table", "one_db_column", "one_db_profile"}:
        if db_names:
            info("Configured DB profiles: " + ", ".join(db_names))
        profile = click.prompt("DB profile name").strip()
        params["db_profile"] = profile
        if req == "one_db_profile":
            assets.append(AssetRef(kind="db_profile", ref=profile))
        else:
            database = click.prompt("Database name", default="", show_default=False).strip()
            schema = click.prompt("Schema name").strip()
            table = click.prompt("Table name").strip()
            params["table"] = f"{schema}.{table}" if schema else table
            if req == "one_db_column":
                column = click.prompt("Column name").strip()
                params["column"] = column
                ref = "/".join(p for p in (profile, database, schema, f"{table}.{column}") if p)
                assets.append(AssetRef(kind="db_column", ref=ref))
            else:
                ref = "/".join(p for p in (profile, database, schema, table) if p)
                assets.append(AssetRef(kind="db_table", ref=ref))
    elif req == "many_db_profiles":
        if db_names:
            info("Configured DB profiles: " + ", ".join(db_names))
        raw = click.prompt("DB profiles (comma-separated)").strip()
        picks = [p.strip() for p in raw.split(",") if p.strip()]
        params["db_profiles"] = ", ".join(f"`{p}`" for p in picks)
        for name in picks:
            assets.append(AssetRef(kind="db_profile", ref=name))
    elif req in {
        "one_asset_notebook",
        "one_asset_job",
        "one_asset_pipeline",
        "one_asset_query",
    }:
        kind = "asset_" + req.removeprefix("one_asset_")
        ref, profile, display = _pick_remote_asset(cfg, kind)
        if ref is not None:
            assets.append(ref)
        params["db_profile"] = profile or "<profile>"
        placeholder = kind.removeprefix("asset_")
        params[placeholder] = display or "<unknown>"
    elif req == "one_lineage":
        rows = _list_lineage_artifacts_safe()
        if rows:
            info(
                "Lineage artifacts: "
                + ", ".join(str(r.get("name") or r.get("id")) for r in rows[:20])
            )
        lineage_ref = click.prompt("Lineage artifact name or id").strip()
        params["lineage"] = lineage_ref
        assets.append(AssetRef(kind="lineage_artifact", ref=lineage_ref))
    elif req == "any":
        assets.extend(_collect_assets_wizard(cfg))

    # Doc-profile RAG retrieval helps virtually every page; offer it
    # regardless of the template's required assets.
    doc_names = sorted((cfg.doc_profiles or {}).keys())
    if doc_names:
        info("Configured doc profiles: " + ", ".join(doc_names))
        pick_doc = click.prompt(
            "Attach doc profile(s)? Comma-separated names (or empty to skip)",
            default="",
            show_default=False,
        ).strip()
        if pick_doc:
            for name in [p.strip() for p in pick_doc.split(",") if p.strip()]:
                assets.append(AssetRef(kind="doc_profile", ref=name))

    return assets, params


def _pick_remote_asset(
    cfg: AMXConfig,
    kind: str,
) -> tuple[AssetRef | None, str, str]:
    """Wizard: pick one ingested remote asset from the active history store.

    Returns ``(ref, profile_name, display_name)``. ``ref`` is ``None``
    if the user skips or no rows exist for any profile.
    """
    spec = _REMOTE_ASSET_TABLES.get(kind)
    if spec is None:
        return None, "", ""
    _, sql = spec

    db_names = sorted((cfg.db_profiles or {}).keys())
    if not db_names:
        info("No DB profiles configured. Run /db add first.")
        return None, "", ""
    info("Configured DB profiles: " + ", ".join(db_names))
    profile = click.prompt("DB profile to pick the asset from").strip()
    if not profile:
        return None, "", ""

    rows = _list_remote_assets_safe(sql, profile)
    if not rows:
        warn(
            f"No ingested {kind.removeprefix('asset_')}s for profile "
            f"`{profile}`. Run /db ingest-assets first."
        )
        return None, profile, ""

    info(f"Ingested {kind.removeprefix('asset_')}s for `{profile}`:")
    for idx, row in enumerate(rows[:30], start=1):
        label_bits = [str(row.get("name") or row.get("id") or "?")]
        loc = row.get("workspace_path") or row.get("qualified_name")
        if loc:
            label_bits.append(f"({loc})")
        click.echo(f"  {idx}. {' '.join(label_bits)}")
    raw = click.prompt("Selection (number)").strip()
    try:
        choice = int(raw)
    except ValueError:
        error(f"Not a number: {raw}")
        return None, profile, ""
    if not 1 <= choice <= len(rows[:30]):
        error(f"Selection out of range: {raw}")
        return None, profile, ""
    chosen = rows[choice - 1]
    asset_id = str(chosen.get("id") or "")
    display = str(chosen.get("name") or chosen.get("qualified_name") or asset_id)
    if not asset_id:
        return None, profile, display
    return AssetRef(kind=kind, ref=f"{profile}:{asset_id}"), profile, display  # type: ignore[arg-type]


def _list_remote_assets_safe(sql: str, profile: str) -> list[dict[str, Any]]:
    try:
        from amx.storage.sqlite_store import history_store

        hs = history_store()
        if hs is None:
            return []
        with hs._connect() as conn:
            cursor = conn.execute(sql, (profile,))
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, row, strict=False)) for row in cursor.fetchall()]
    except Exception as exc:  # noqa: BLE001
        warn(f"Could not list ingested assets: {exc}")
        return []


def _list_lineage_artifacts_safe() -> list[dict[str, Any]]:
    try:
        from amx.lineage.store import list_lineage_artifacts
        from amx.storage.sqlite_store import history_store

        hs = history_store()
        return list_lineage_artifacts(hs) if hs is not None else []
    except Exception:  # noqa: BLE001
        return []


def _collect_assets_wizard(cfg: AMXConfig) -> list[AssetRef]:
    """Walk the user through DB / doc-profile / lineage pickers."""
    assets: list[AssetRef] = []

    # DB profile assets — list configured profiles, let the user pick by name.
    db_names = sorted((cfg.db_profiles or {}).keys())
    if db_names:
        info("Configured DB profiles: " + ", ".join(db_names))
    pick_db = click.prompt(
        "Attach DB profile(s)? Comma-separated names (or empty to skip)",
        default="",
        show_default=False,
    ).strip()
    if pick_db:
        for name in [p.strip() for p in pick_db.split(",") if p.strip()]:
            assets.append(AssetRef(kind="db_profile", ref=name))

    # Doc profile assets.
    doc_names = sorted((cfg.doc_profiles or {}).keys())
    if doc_names:
        info("Configured doc profiles: " + ", ".join(doc_names))
    pick_doc = click.prompt(
        "Attach doc profile(s)? Comma-separated names (or empty to skip)",
        default="",
        show_default=False,
    ).strip()
    if pick_doc:
        for name in [p.strip() for p in pick_doc.split(",") if p.strip()]:
            assets.append(AssetRef(kind="doc_profile", ref=name))

    # Lineage artifacts — read from the history store.
    artifact_rows = _list_lineage_artifacts_safe()
    if artifact_rows:
        info(
            "Lineage artifacts: "
            + ", ".join(str(r.get("name") or r.get("id")) for r in artifact_rows[:20])
        )
    pick_lin = click.prompt(
        "Attach lineage artifact(s)? Comma-separated names/ids (or empty to skip)",
        default="",
        show_default=False,
    ).strip()
    if pick_lin:
        for ref in [p.strip() for p in pick_lin.split(",") if p.strip()]:
            assets.append(AssetRef(kind="lineage_artifact", ref=ref))

    return assets


def _collect_sources_wizard() -> list[Path]:
    """Prompt for local file paths to attach as sources.

    Sources are scratch input for THIS page only. For permanent RAG
    storage, register a doc profile via ``/docs add`` and attach it as
    an asset; the page generator will RAG-query the indexed corpus
    automatically. This wizard step deliberately defaults to empty.
    """
    raw = click.prompt(
        "Attach ad-hoc source files for THIS page? Comma-separated paths "
        "(empty to skip; for permanent indexing use /docs add instead)",
        default="",
        show_default=False,
    ).strip()
    if not raw:
        return []
    paths: list[Path] = []
    for chunk in raw.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        path = Path(chunk).expanduser()
        if not path.is_file():
            warn(f"Skipping {path}: not a regular file")
            continue
        _notice_multi_sheet_xlsx(path)
        paths.append(path)
    return paths


def _notice_multi_sheet_xlsx(path: Path) -> None:
    """Tell the user how many sheets a source xlsx has so they aren't
    surprised when the LLM sees one big concatenated markdown rather
    than just the first sheet."""
    if path.suffix.lower() != ".xlsx":
        return
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception:  # noqa: BLE001
        return
    if len(sheets) > 1:
        info(
            f"note: {path.name} has {len(sheets)} sheets — all will be "
            "inlined as separate tables in the page context."
        )


def _maybe_promote_to_doc_profile(cfg: AMXConfig, paths: list[Path]) -> None:
    """Offer to also index source paths into a permanent doc profile.

    Reuses the same upload + ingest path as ``/docs add`` so the file
    becomes RAG-retrievable for future pages.
    """
    if not paths:
        return
    doc_names = sorted((cfg.doc_profiles or {}).keys())
    if not doc_names:
        return
    if not click.confirm(
        "Also index these source(s) into a doc profile for future pages?",
        default=False,
    ):
        return
    info("Doc profiles: " + ", ".join(doc_names))
    profile = click.prompt("Target doc profile name").strip()
    if not profile:
        return
    try:
        from amx.docs.rag import RAGStore
        from amx.docs.scanner import cleanup_scan_artifacts, scan_all_sources
        from amx.docs.uploads import UploadError, save_uploaded_file
    except ImportError as exc:
        error(f"docs subsystem unavailable: {exc}")
        return

    saved: list[Any] = []
    for path in paths:
        try:
            payload = path.read_bytes()
            res = save_uploaded_file(cfg, profile, path.name, payload)
            saved.append(res)
        except UploadError as exc:
            error(f"{path}: {exc}")
        except OSError as exc:
            error(f"{path}: {exc}")
    if not saved:
        return
    upload_root = str(Path(saved[0].saved_path).parent)
    documents: list[Any] = []
    try:
        scan_result = scan_all_sources([upload_root])
        documents = list(scan_result)
        store = RAGStore()
        summary = store.ingest(documents, refresh=False)
        success(
            f"Indexed {len(saved)} file(s) into doc profile `{profile}` "
            f"({summary.chunk_count} chunks)"
        )
    except Exception as exc:  # noqa: BLE001
        error(f"Ingest failed: {exc}. Files saved; run `/docs ingest` to retry.")
    finally:
        cleanup_scan_artifacts(documents)


def _attach_sources(
    cfg: AMXConfig,
    page_id: str,
    paths: list[Path],
) -> list[SourceRef]:
    """Copy each path into the per-page upload root and return SourceRefs."""
    if not paths:
        return []
    from amx.docs.uploads import UploadError, save_uploaded_file

    refs: list[SourceRef] = []
    profile = f"pages/{page_id}"
    for path in paths:
        try:
            payload = path.read_bytes()
            res = save_uploaded_file(cfg, profile, path.name, payload)
        except UploadError as exc:
            error(f"{path}: {exc}")
            continue
        except OSError as exc:
            error(f"{path}: {exc}")
            continue
        refs.append(SourceRef(kind="upload", path=res.saved_path, original_name=res.original_name))
        info(f"Attached source: {path.name}")
    return refs


def _hard_delete_page(page_id: str) -> bool:
    """Hard-delete a page and its child rows. Returns True on success."""
    from amx.storage.sqlite_store import history_store

    hs = history_store()
    if hs is None:
        return False
    try:
        with hs._connect() as conn:
            conn.execute("DELETE FROM documentation_page_versions WHERE page_id = ?", (page_id,))
            conn.execute("DELETE FROM documentation_page_sources WHERE page_id = ?", (page_id,))
            conn.execute("DELETE FROM documentation_page_assets WHERE page_id = ?", (page_id,))
            conn.execute("DELETE FROM documentation_pages WHERE id = ?", (page_id,))
    except Exception as exc:  # noqa: BLE001
        error(f"Hard delete failed: {exc}")
        return False
    return True


def register_pages_commands(
    main: click.Group,
    *,
    finalize_scope: FinalizeScope | None = None,
) -> click.Group:
    """Attach the ``/pages`` namespace + subcommands to the main Click group."""
    del finalize_scope  # reserved for future per-asset scope expansion

    @main.group()
    def pages() -> None:
        """Compose, edit, and export documentation pages."""

    @pages.command("new")
    @click.option("--title", default=None, help="Page title (prompted when omitted).")
    @click.option("--intent", default=None, help="Free-text generation intent.")
    @click.option(
        "--intent-template",
        "intent_template_slug",
        default=None,
        help=(
            "Preset intent shape. Slugs: "
            + ", ".join(t.slug for t in INTENT_TEMPLATES)
            + ". Bypassed when --intent is also given."
        ),
    )
    @click.option(
        "--asset",
        "assets_flag",
        multiple=True,
        help="Repeatable. Format: KIND:REF (e.g. db_profile:prod, doc_profile:eng).",
    )
    @click.option(
        "--source",
        "sources_flag",
        multiple=True,
        type=click.Path(exists=True, dir_okay=False, path_type=Path),
        help=(
            "Repeatable. Local file path to attach as a source for THIS page "
            "only. For permanent RAG indexing, use /docs add instead and "
            "attach the doc profile as an asset."
        ),
    )
    @click.option(
        "--no-generate",
        is_flag=True,
        default=False,
        help="Create the draft but skip the immediate LLM composition.",
    )
    @click.pass_obj
    def pages_new(
        cfg: AMXConfig,
        title: str | None,
        intent: str | None,
        intent_template_slug: str | None,
        assets_flag: tuple[str, ...],
        sources_flag: tuple[Path, ...],
        no_generate: bool,
    ) -> None:
        """Create a new documentation page (wizard when flags are omitted)."""
        svc = _svc(cfg)

        # Title: flag wins, otherwise prompt.
        final_title = (title or "").strip()
        if not final_title:
            final_title = click.prompt("Page title").strip()
            if not final_title:
                error("Title is required.")
                return

        # Intent template + assets: resolve in one step when no
        # asset/intent flags were given, since the template choice
        # drives which assets the wizard collects.
        final_assets: list[AssetRef]
        final_intent: str = (intent or "").strip()
        template_params: dict[str, str] = {}

        if assets_flag:
            try:
                final_assets = [_parse_asset_flag(a) for a in assets_flag]
            except click.BadParameter as exc:
                error(str(exc))
                return
            # Power-user mode: assets came from flags. If the user also
            # passed --intent-template, render it with no params; the
            # user can supply --intent to override.
            if not final_intent and intent_template_slug:
                tpl = template_by_slug(intent_template_slug)
                if tpl is None:
                    error(f"Unknown intent template: {intent_template_slug}")
                    return
                final_intent = render_intent(tpl)
        else:
            # Wizard mode.
            template: IntentTemplate | None = None
            if intent_template_slug:
                template = template_by_slug(intent_template_slug)
                if template is None:
                    error(f"Unknown intent template: {intent_template_slug}")
                    return
            elif not final_intent:
                template = _pick_intent_template()

            if template is not None:
                final_assets, template_params = _collect_template_params(cfg, template)
                if not final_intent:
                    final_intent = render_intent(template, **template_params)
            else:
                final_assets = _collect_assets_wizard(cfg)

        # Intent fallback for the "Custom" picker choice.
        if not final_intent:
            final_intent = click.prompt("Intent (free text)", default="").strip()

        # Sources: flag wins; otherwise prompt for paths.
        source_paths = list(sources_flag) if sources_flag else _collect_sources_wizard()
        if source_paths and not sources_flag:
            _maybe_promote_to_doc_profile(cfg, source_paths)

        now = _utcnow()
        page_id = svc.create_draft(
            title=final_title,
            intent=final_intent,
            assets=final_assets,
            sources=[],
            created_by=None,
            now=now,
        )

        attached = _attach_sources(cfg, page_id, source_paths)
        for ref in attached:
            svc.store.history.attach_documentation_page_source(
                page_id,
                source_kind=ref.kind,
                source_path=ref.path,
                original_name=ref.original_name,
                created_at=now,
            )

        success(f"Created page {page_id}")

        if no_generate:
            info("Skipped LLM composition (--no-generate). Use /pages edit to write the body.")
            return

        try:
            svc.generate(page_id, now=_utcnow())
        except Exception as exc:  # noqa: BLE001
            error(f"Generation failed: {exc}. Draft was saved; re-run with /pages edit.")
            return

        page = svc.store.get(page_id)
        body = (page or {}).get("markdown_body", "") or ""
        preview = "\n".join(body.splitlines()[:10])
        info("Preview (first 10 lines):")
        click.echo(preview)

    @pages.command("list")
    @click.pass_obj
    def pages_list(cfg: AMXConfig) -> None:
        """List active documentation pages."""
        svc = _svc(cfg)
        rows = svc.store.list_active()
        if not rows:
            info("No pages yet. Create one with /pages new.")
            return
        table = Table(title="Documentation pages", show_lines=False)
        table.add_column("ID")
        table.add_column("Title")
        table.add_column("Status")
        table.add_column("Updated")
        for r in rows:
            table.add_row(
                str(r.get("id", "")),
                str(r.get("title", "")),
                str(r.get("status", "")),
                str(r.get("updated_at", "")),
            )
        Console().print(table)

    @pages.command("show")
    @click.argument("page_id")
    @click.pass_obj
    def pages_show(cfg: AMXConfig, page_id: str) -> None:
        """Print the markdown body of a page to stdout."""
        svc = _svc(cfg)
        page = svc.store.get(page_id)
        if page is None:
            error(f"Page {page_id} not found.")
            return
        click.echo(page.get("markdown_body", "") or "")

    @pages.command("edit")
    @click.argument("page_id")
    @click.option("--note", default=None, help="Optional revision note.")
    @click.pass_obj
    def pages_edit(cfg: AMXConfig, page_id: str, note: str | None) -> None:
        """Open the page body in $EDITOR and save the result as a new revision."""
        svc = _svc(cfg)
        page = svc.store.get(page_id)
        if page is None:
            error(f"Page {page_id} not found.")
            return
        original = page.get("markdown_body", "") or ""

        import tempfile

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".md", delete=False, encoding="utf-8"
        ) as fh:
            fh.write(original)
            tmp_path = Path(fh.name)
        try:
            click.edit(filename=str(tmp_path), require_save=False, extension=".md")
            edited = tmp_path.read_text(encoding="utf-8")
        finally:
            try:
                tmp_path.unlink()
            except OSError:
                pass

        if edited == original:
            info("No changes — revision not saved.")
            return

        svc.save_revision(
            page_id,
            markdown_body=edited,
            now=_utcnow(),
            saved_by=None,
            note=note,
        )
        success(f"Saved new revision for {page_id}")

    @pages.command("export")
    @click.argument("page_id")
    @click.option(
        "--format",
        "fmt",
        type=click.Choice(["md", "pdf"]),
        required=True,
        help="Export format.",
    )
    @click.option(
        "--out",
        "out_path",
        type=click.Path(dir_okay=False, path_type=Path),
        default=None,
        help="Output file path (stdout when omitted).",
    )
    @click.pass_obj
    def pages_export(
        cfg: AMXConfig,
        page_id: str,
        fmt: str,
        out_path: Path | None,
    ) -> None:
        """Export a page as Markdown or PDF."""
        svc = _svc(cfg)
        try:
            payload = svc.export(page_id, fmt)  # type: ignore[arg-type]
        except KeyError:
            error(f"Page {page_id} not found.")
            return

        if out_path is not None:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            if isinstance(payload, bytes):
                out_path.write_bytes(payload)
            else:
                out_path.write_text(payload, encoding="utf-8")
            success(f"Exported page {page_id} -> {out_path}")
            return

        if isinstance(payload, bytes):
            sys.stdout.buffer.write(payload)
        else:
            click.echo(payload, nl=False)

    @pages.command("delete")
    @click.argument("page_id")
    @click.option(
        "--purge",
        is_flag=True,
        default=False,
        help="Hard-delete the page and all related rows.",
    )
    @click.pass_obj
    def pages_delete(cfg: AMXConfig, page_id: str, purge: bool) -> None:
        """Soft-delete a page (use --purge to hard-delete)."""
        svc = _svc(cfg)
        page = svc.store.get(page_id)
        if page is None:
            error(f"Page {page_id} not found.")
            return

        if purge:
            if not _hard_delete_page(page_id):
                error(f"Could not purge {page_id}.")
                return
            success(f"Purged page {page_id}")
            return

        svc.soft_delete(page_id, now=_utcnow())
        success(f"Soft-deleted page {page_id}")

    @pages.command("assign-profile")
    @click.argument("slug", required=False, default=None)
    @click.option(
        "--profile",
        "db_profile",
        default=None,
        help="DB profile name to associate with the page (omit to clear).",
    )
    @click.pass_obj
    def pages_assign_profile(
        cfg: AMXConfig,
        slug: str | None,
        db_profile: str | None,
    ) -> None:
        """Associate a db_profile with a page for team-scoped filtering.

        When called without arguments, runs an interactive wizard:
          1. Pick a page from the list of active pages.
          2. Pick a DB profile from the configured profiles.
          3. Confirm and apply.

        Power-user shortcut (non-interactive)::

            /pages assign-profile <slug> --profile <name>

        Pass ``--profile ""`` (empty string) or omit ``--profile`` in the
        wizard to clear the field, marking the page as unscoped /
        cross-profile.
        """
        svc = _svc(cfg)

        # ── Wizard path ───────────────────────────────────────────────────
        if slug is None or db_profile is None:
            # Step 1: page picker when slug not supplied.
            final_slug = slug
            if final_slug is None:
                active = svc.store.list_active()
                if not active:
                    info("No active pages found. Create one with /pages new.")
                    return
                info("Active pages:")
                for idx, row in enumerate(active, start=1):
                    click.echo(f"  {idx}. [{row.get('slug', '')}] {row.get('title', '')}")
                raw = click.prompt("Select page (number or slug)").strip()
                try:
                    choice = int(raw)
                    if 1 <= choice <= len(active):
                        final_slug = str(active[choice - 1]["slug"])
                    else:
                        error(f"Selection out of range: {raw}")
                        return
                except ValueError:
                    final_slug = raw

            # Step 2: profile picker when --profile not supplied.
            final_profile: str | None = db_profile
            if final_profile is None:
                db_names = sorted((cfg.db_profiles or {}).keys())
                if db_names:
                    info("Configured DB profiles: " + ", ".join(db_names))
                raw_profile = click.prompt(
                    "DB profile to assign (empty to clear / mark unscoped)",
                    default="",
                    show_default=False,
                ).strip()
                final_profile = raw_profile if raw_profile else None

            # Step 3: confirm.
            profile_label = f"'{final_profile}'" if final_profile else "(clear / unscoped)"
            if not click.confirm(
                f"Assign profile {profile_label} to page '{final_slug}'?", default=True
            ):
                info("Cancelled.")
                return
        else:
            final_slug = slug
            final_profile = db_profile if db_profile else None

        # ── Apply ─────────────────────────────────────────────────────────
        updated = svc.store.assign_db_profile(
            slug=final_slug,
            db_profile=final_profile,
            now=_utcnow(),
        )
        if not updated:
            error(f"No page with slug '{final_slug}' found.")
            return
        if final_profile:
            success(f"Page '{final_slug}' is now scoped to profile '{final_profile}'.")
        else:
            success(f"Page '{final_slug}' is now unscoped (db_profile cleared).")

    return pages


__all__ = ["register_pages_commands"]
