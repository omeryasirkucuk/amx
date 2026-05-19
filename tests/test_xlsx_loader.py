"""Tests for the xlsx loader used by the docs scanner and pages module."""

from __future__ import annotations

from pathlib import Path

import openpyxl


def test_xlsx_loader_serialises_each_sheet_as_markdown_table(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    p = tmp_path / "sample.xlsx"
    wb.save(p)

    from amx.docs.loaders.xlsx_loader import load_xlsx

    text = load_xlsx(p)
    assert "## Customers" in text
    assert "| id | name |" in text
    assert "| 1 | Alice |" in text
    assert "| 2 | Bob |" in text


def test_xlsx_loader_skips_empty_sheets(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Filled"
    wb.active.append(["a", "b"])
    wb.create_sheet("Empty")
    p = tmp_path / "mix.xlsx"
    wb.save(p)

    from amx.docs.loaders.xlsx_loader import load_xlsx

    text = load_xlsx(p)
    assert "## Filled" in text
    assert "## Empty" not in text


def test_xlsx_loader_truncates_long_cells(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Big"
    ws.append(["x"])
    ws.append(["a" * 10_000])
    p = tmp_path / "big.xlsx"
    wb.save(p)

    from amx.docs.loaders.xlsx_loader import MAX_CELL_LEN, load_xlsx

    text = load_xlsx(p)
    assert "…" in text
    # No emitted cell exceeds the limit.
    for line in text.splitlines():
        for cell in line.split("|"):
            assert len(cell.strip()) <= MAX_CELL_LEN
